from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

from core.json_utils import parse_json_result
from core.llm_client import ChatMessage, LLMClient
from core.stage_base import PromptStage
from core.state import ProjectState
from core.storage import ProjectStorage


HEADERS = [
    "场次",
    "镜号",
    "镜头类型/景别",
    "画面内容",
    "人物调度",
    "摄影机角度",
    "摄影机运动",
    "声音/对白",
    "剪辑/转场",
    "时长",
    "镜头目的",
]

STORYBOARD_CHUNK_MAX_CHARS = 600
STORYBOARD_MAX_SHOTS_PER_BATCH = 6
STORYBOARD_MAX_DIALOGUES_PER_CHUNK = 3
STORYBOARD_BATCH_RETRIES = 2
STORYBOARD_MIN_SPLIT_CHARS = 180

HEADER_ALIASES = {
    "场次": {"场次", "章节", "章", "场"},
    "镜号": {"镜号", "镜头号", "镜头编号"},
    "镜头类型/景别": {"镜头类型/景别", "景别/镜头", "景别", "镜头类型"},
    "画面内容": {"画面内容", "画面", "镜头内容"},
    "人物调度": {"人物调度", "调度"},
    "摄影机角度": {"摄影机角度", "机位角度", "角度"},
    "摄影机运动": {"摄影机运动", "摄影机移动", "运动"},
    "声音/对白": {"声音/对白", "声音", "对白"},
    "剪辑/转场": {"剪辑/转场", "剪辑", "转场"},
    "时长": {"时长", "镜头时长", "预计时长"},
    "镜头目的": {"镜头目的", "目的", "镜头任务"},
}


class StoryboardXlsxStage(PromptStage):
    def run(self, state: ProjectState) -> ProjectState:
        content = generate_chunked_storyboard_csv(state, self.llm)
        saved_path = self.save_content(state, content)

        stage_files = dict(state.get("stage_files", {}))
        stage_files[self.output_key] = saved_path
        return {
            self.output_key: content,
            "stage_files": stage_files,
        }

    def save_content(self, state: ProjectState, content: str) -> str:
        output_path = Path(state["output_dir"]) / self.output_filename
        write_storyboard_xlsx(content, output_path)
        return str(output_path.resolve())


def make_storyboard_stage(prompt_dir: Path, llm: LLMClient, storage: ProjectStorage) -> PromptStage:
    return StoryboardXlsxStage(
        name="storyboard",
        prompt_file="storyboard_prompt.txt",
        output_key="storyboard",
        output_filename="06_storyboard.xlsx",
        prompt_dir=prompt_dir,
        llm=llm,
        storage=storage,
        placeholders={"scene_write": "final_script", "duration_minutes": "duration_minutes"},
        system_prompt="你是一名从业数十年的顶级电影导演。严格遵守用户给定提示词，不改变原意。",
    )


def generate_chunked_storyboard_csv(state: ProjectState, llm: LLMClient) -> str:
    script_text = str(state.get("final_script") or "").strip()
    if not script_text:
        raise ValueError("final_script is required before generating storyboard.")

    chunks = split_script_for_storyboards(script_text)
    rows: list[list[str]] = [HEADERS]
    for scene_index, chunk in enumerate(chunks, start=1):
        frames = generate_storyboard_batch(
            llm,
            chunk,
            scene_index=scene_index,
            state=state,
        )
        rows.extend(storyboard_frames_to_rows(frames, scene_index=scene_index))
        if scene_index < len(chunks):
            rows.append([])
    return rows_to_csv(rows)


def generate_storyboard_batch(
    llm: LLMClient,
    chunk: str,
    *,
    scene_index: int,
    state: ProjectState,
) -> list[dict[str, Any]]:
    last_error = ""
    token_callback = state.get("llm_token_callback")
    for attempt in range(1, STORYBOARD_BATCH_RETRIES + 1):
        messages = [
            ChatMessage(
                role="system",
                content="你是分镜导演。只输出合法 JSON array，不输出 Markdown、解释或多余文本。",
            ),
            ChatMessage(role="user", content=_storyboard_batch_prompt(chunk, scene_index, state, last_error)),
        ]
        if callable(token_callback):
            raw = llm.generate_stream(
                messages,
                stage_name="storyboard_chunk",
                on_token=lambda token, batch_index=scene_index: token_callback(
                    "storyboard_chunk",
                    "storyboard",
                    token,
                    {"batch_index": batch_index},
                ),
            )
        else:
            raw = llm.generate(messages, stage_name="storyboard_chunk")
        frames = parse_json_result(raw, expected_type="list")
        if isinstance(frames, list) and frames:
            normalized = [frame if isinstance(frame, dict) else {"画面内容": str(frame)} for frame in frames]
            if len(normalized) > STORYBOARD_MAX_SHOTS_PER_BATCH:
                last_error = f"当前批次返回 {len(normalized)} 个镜头，超过上限 {STORYBOARD_MAX_SHOTS_PER_BATCH}。"
                continue
            missing = validate_storyboard_coverage(chunk, normalized)
            if not missing:
                return normalized
            last_error = "缺少对白覆盖：" + "；".join(missing[:5])
        else:
            last_error = f"未返回合法 JSON array，可能被截断。模型输出片段：{raw[:300]}"

    sub_chunks = split_script_for_storyboards(
        chunk,
        max_chars=max(STORYBOARD_MIN_SPLIT_CHARS, len(chunk) // 2),
        max_dialogues=max(1, STORYBOARD_MAX_DIALOGUES_PER_CHUNK - 1),
    )
    if len(sub_chunks) > 1:
        frames: list[dict[str, Any]] = []
        for offset, sub_chunk in enumerate(sub_chunks, start=1):
            frames.extend(
                generate_storyboard_batch(
                    llm,
                    sub_chunk,
                    scene_index=scene_index * 100 + offset,
                    state=state,
                )
            )
        return frames

    missing = validate_storyboard_coverage(chunk, [])
    raise ValueError(
        "分镜批次生成失败，已重试并尝试拆分但仍无法得到完整结果。"
        f"场次：{scene_index}；错误：{last_error}；缺失内容：{missing[:5]}；文本片段：{chunk[:300]}"
    )


def split_script_for_storyboards(
    script_text: str,
    *,
    max_chars: int = STORYBOARD_CHUNK_MAX_CHARS,
    max_dialogues: int = STORYBOARD_MAX_DIALOGUES_PER_CHUNK,
) -> list[str]:
    units: list[str] = []
    for block in re.split(r"\n\s*\n", script_text.strip()):
        block = block.strip()
        if not block:
            continue
        if len(block) <= max_chars:
            units.append(block)
        else:
            units.extend(_split_long_text(block, max_chars))

    chunks: list[str] = []
    current: list[str] = []
    current_len = 0
    current_dialogues = 0
    for unit in units:
        unit_dialogues = len(extract_dialogue_lines(unit))
        would_exceed = (
            current
            and (current_len + len(unit) > max_chars or current_dialogues + unit_dialogues > max_dialogues)
        )
        if would_exceed:
            chunks.append("\n\n".join(current).strip())
            current = []
            current_len = 0
            current_dialogues = 0
        current.append(unit)
        current_len += len(unit)
        current_dialogues += unit_dialogues
    if current:
        chunks.append("\n\n".join(current).strip())
    return chunks or [script_text.strip()]


def extract_dialogue_lines(text: str) -> list[str]:
    excluded = {"场景", "地点", "时间", "音效", "声音", "备注", "转场", "画面", "预计时长", "本场预计时长", "连续性备注"}
    dialogues: list[str] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or len(line) > 220:
            continue
        match = re.match(r"^([\u4e00-\u9fffA-Za-z0-9_·（）()]{1,30})\s*[:：]\s*(.+)$", line)
        if not match:
            continue
        speaker = match.group(1).strip()
        content = match.group(2).strip()
        if speaker in excluded or not content:
            continue
        dialogues.append(content)
    return dialogues


def validate_storyboard_coverage(chunk: str, frames: list[dict[str, Any]]) -> list[str]:
    required = extract_dialogue_lines(chunk)
    if not required:
        return []

    covered_text = "\n".join(
        str(frame.get("台词") or frame.get("声音/对白") or frame.get("dialogue") or frame.get("sound") or "")
        for frame in frames
    )
    normalized_covered = _normalize_dialogue(covered_text)
    return [dialogue for dialogue in required if _normalize_dialogue(dialogue) not in normalized_covered]


def storyboard_frames_to_rows(frames: list[dict[str, Any]], *, scene_index: int) -> list[list[str]]:
    rows: list[list[str]] = []
    for shot_index, frame in enumerate(frames, start=1):
        sound = _frame_value(frame, "声音/对白", "台词", "dialogue", "sound")
        rows.append(
            [
                str(scene_index),
                f"{scene_index}-{shot_index}",
                _frame_value(frame, "镜头类型/景别", "景别", "shot_type"),
                _frame_value(frame, "画面内容", "画面", "image", "visual"),
                _frame_value(frame, "人物调度", "调度", "blocking"),
                _frame_value(frame, "摄影机角度", "角度", "camera_angle"),
                _frame_value(frame, "摄影机运动", "运动", "camera_movement"),
                sound or "无",
                _frame_value(frame, "剪辑/转场", "转场", "edit"),
                _frame_value(frame, "时长", "duration") or "6秒",
                _frame_value(frame, "镜头目的", "目的", "purpose"),
            ]
        )
    return rows


def rows_to_csv(rows: list[list[str]]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerows(rows)
    return buffer.getvalue()


def _storyboard_batch_prompt(chunk: str, scene_index: int, state: ProjectState, last_error: str = "") -> str:
    error_part = f"\n上一次失败原因：{last_error}\n请修正。" if last_error else ""
    return f"""
只把下面这个剧本文本批次转成分镜，不要处理批次外内容。
每批最多输出 {STORYBOARD_MAX_SHOTS_PER_BATCH} 个镜头。
所有真实对白必须出现在“台词”或“声音/对白”字段；没有对白写“无”。
场景描述、音效、备注不要误写成对白。
输出只能是一个 JSON 数组，数组元素为对象。

影片目标时长：{state.get('duration_minutes') or '未指定'}分钟
批次编号：{scene_index}
{error_part}

剧本文本批次：
{chunk}

JSON 字段：
[
  {{
    "镜头类型/景别": "中景",
    "画面内容": "...",
    "人物调度": "...",
    "摄影机角度": "...",
    "摄影机运动": "...",
    "声音/对白": "...",
    "剪辑/转场": "...",
    "时长": "6秒",
    "镜头目的": "..."
  }}
]
""".strip()


def _split_long_text(text: str, max_chars: int) -> list[str]:
    parts = [part.strip() for part in re.split(r"(?<=[。！？!?])", text) if part.strip()]
    if not parts:
        return [text[index : index + max_chars] for index in range(0, len(text), max_chars)]

    chunks: list[str] = []
    current = ""
    for part in parts:
        if len(part) > max_chars:
            if current:
                chunks.append(current.strip())
                current = ""
            chunks.extend(part[index : index + max_chars] for index in range(0, len(part), max_chars))
            continue
        if current and len(current) + len(part) > max_chars:
            chunks.append(current.strip())
            current = ""
        current += part
    if current:
        chunks.append(current.strip())
    return chunks


def _frame_value(frame: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = frame.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _normalize_dialogue(text: str) -> str:
    return re.sub(r"\s+", "", text).replace("“", "").replace("”", "").replace('"', "")


def write_storyboard_xlsx(csv_content: str, output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on local env
        raise RuntimeError("Package 'openpyxl' is missing. Run: pip install -r requirements.txt") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = normalize_storyboard_rows(csv_content)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "分镜表"
    worksheet.sheet_view.showGridLines = False

    for row in rows:
        worksheet.append(row if row else [""] * len(HEADERS))

    thin_side = Side(style="thin", color="B8C2CC")
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill("solid", fgColor="E7EBF0")
    body_fill = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(name="Microsoft YaHei", bold=True, color="44505C", size=11)
    body_font = Font(name="Microsoft YaHei", color="5F6B7A", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    widths = [12, 10, 18, 42, 30, 18, 18, 32, 22, 12, 34]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    blank_rows = {
        row_index
        for row_index in range(2, worksheet.max_row + 1)
        if all((worksheet.cell(row=row_index, column=column).value or "") == "" for column in range(1, len(HEADERS) + 1))
    }

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=len(HEADERS)):
        row_index = row[0].row
        is_header = row_index == 1
        is_blank = row_index in blank_rows
        worksheet.row_dimensions[row_index].height = 10 if is_blank else (28 if is_header else 58)

        for cell in row:
            if is_blank:
                cell.border = Border()
                cell.fill = body_fill
                continue

            cell.border = grid_border
            cell.fill = header_fill if is_header else body_fill
            cell.font = header_font if is_header else body_font
            cell.alignment = header_alignment if is_header else body_alignment

    worksheet.freeze_panes = "A2"
    workbook.save(output_path)


def normalize_storyboard_rows(content: str) -> list[list[str]]:
    rows = parse_csv_rows(strip_code_fence(content))
    if not rows:
        return [HEADERS]

    first_nonblank = next((row for row in rows if row), None)
    if first_nonblank is None:
        return [HEADERS]

    normalized: list[list[str]] = []
    if is_header_row(first_nonblank):
        header_consumed = False
        for row in rows:
            if row and not header_consumed and is_header_row(row):
                normalized.append(HEADERS)
                header_consumed = True
                continue
            normalized.append(normalize_data_row(row) if row else [])
    else:
        normalized.append(HEADERS)
        normalized.extend(normalize_data_row(row) if row else [] for row in rows)

    return normalized


def parse_csv_rows(content: str) -> list[list[str]]:
    reader = csv.reader(io.StringIO(content))
    rows: list[list[str]] = []
    for row in reader:
        cleaned = [cell.strip() for cell in row]
        if not cleaned or all(cell == "" for cell in cleaned):
            rows.append([])
        else:
            rows.append(cleaned)
    return rows


def normalize_data_row(row: list[str]) -> list[str]:
    if len(row) == len(HEADERS) - 1:
        return normalize_row(row[:9] + [""] + row[9:])
    return normalize_row(row)


def normalize_row(row: list[str]) -> list[str]:
    values = row[: len(HEADERS)]
    if len(values) < len(HEADERS):
        values.extend([""] * (len(HEADERS) - len(values)))
    return values


def is_header_row(row: list[str]) -> bool:
    normalized_cells = {normalize_header_cell(cell) for cell in row if cell.strip()}
    if not normalized_cells:
        return False

    matches = 0
    for header, aliases in HEADER_ALIASES.items():
        possible = {normalize_header_cell(value) for value in aliases | {header}}
        if normalized_cells & possible:
            matches += 1
    return matches >= 5


def normalize_header_cell(value: str) -> str:
    return value.strip().replace(" ", "").replace("\t", "").replace("|", "").lower()


def strip_code_fence(content: str) -> str:
    text = content.strip().lstrip("\ufeff")
    if not text.startswith("```"):
        return text

    lines = text.splitlines()
    if lines and lines[0].strip().startswith("```"):
        lines = lines[1:]
    if lines and lines[-1].strip().startswith("```"):
        lines = lines[:-1]
    return "\n".join(lines).strip()
