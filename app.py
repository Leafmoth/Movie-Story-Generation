from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path
from threading import Lock
from typing import Any, Dict, Literal, Optional
from uuid import uuid4

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field

from core.orchestrator import StoryOrchestrator
from core.storage import parse_json_object
from schemas.story import StoryboardGenerationRequest, StoryGenerationRequest, StoryGenerationResponse, response_from_state


app = FastAPI(title="Movie Story Generation", version="0.1.0")
orchestrator = StoryOrchestrator()
app.add_middleware(
    CORSMiddleware,
    allow_origins=list(orchestrator.settings.cors_allow_origins),
    allow_credentials=orchestrator.settings.cors_allow_credentials,
    allow_methods=list(orchestrator.settings.cors_allow_methods),
    allow_headers=list(orchestrator.settings.cors_allow_headers),
)

JobStatus = Literal["pending", "running", "succeeded", "failed"]


class StoryJobCreateResponse(BaseModel):
    job_id: str
    status: JobStatus
    status_url: str
    result_url: str


class StoryJobStatusResponse(BaseModel):
    job_id: str
    status: JobStatus
    created_at: str
    updated_at: str
    project_id: Optional[str] = None
    error: Optional[str] = None
    result_url: str
    files: Dict[str, str] = Field(default_factory=dict)


class StoryJobResultResponse(BaseModel):
    job_id: str
    status: JobStatus
    result: Optional[StoryGenerationResponse] = None
    error: Optional[str] = None


class CharactersResponse(BaseModel):
    job_id: Optional[str] = None
    project_id: Optional[str] = None
    status: Optional[JobStatus] = None
    characters: str
    display_text: str
    parsed: Any = None
    file_url: Optional[str] = None


class StageSectionResponse(BaseModel):
    job_id: Optional[str] = None
    project_id: Optional[str] = None
    status: Optional[JobStatus] = None
    section: str
    content: str
    display_text: str
    parsed: Any = None
    file_url: Optional[str] = None
    relationships: Optional[str] = None
    biography: Optional[str] = None
    outline: Optional[str] = None


class ProjectStoryboardResponse(BaseModel):
    project_id: str
    rows: list[list[str]]
    display_text: str
    file_url: Optional[str] = None


job_executor = ThreadPoolExecutor(max_workers=2)
jobs_lock = Lock()
jobs: dict[str, dict[str, Any]] = {}

FILE_TYPE_ALIASES = {
    "logline": "logline",
    "characters": "characters",
    "character_relations": "character_relations",
    "relationships": "character_relations",
    "biography": "biography",
    "outline": "outline",
    "final_script": "final_script",
    "script": "final_script",
    "storyboard": "storyboard",
    "storyboard_xlsx": "storyboard",
}

FALLBACK_FILENAMES = {
    "logline": "01_logline.json",
    "characters": "02_characters.json",
    "character_relations": "03_relationships.json",
    "biography": "04_biography.json",
    "outline": "05_outline.json",
    "final_script": "final_script.md",
    "storyboard": "06_storyboard.xlsx",
}

STAGE_SECTION_KEYS = {
    "relationships": "character_relations",
    "biography": "biography",
    "outline": "outline",
}

STAGE_FILE_KEYS = {
    "relationships": "character_relations",
    "biography": "biography",
    "outline": "outline",
}

DISPLAY_LABELS = {
    "characters": "角色设置",
    "protagonist": "主角",
    "antagonist": "对手",
    "emotional_core": "情感核心人物",
    "ally": "盟友",
    "mirror": "镜像人物",
    "relationships": "人物关系",
    "name": "关系名称",
    "surface_relation": "表面关系",
    "a_wants_from_b": "A 想从 B 那里得到",
    "b_wants_from_a": "B 想从 A 那里得到",
    "conflict": "冲突",
    "hidden_thing": "隐藏的东西",
    "ending_change": "最后的变化",
    "protagonist_biography": "主角小传",
    "core_supporting_biographies": "核心配角小传",
    "functional_supporting_biographies": "功能性配角",
    "duration_minutes": "影片时长",
    "three_act_outline": "三幕式大纲",
    "act_1": "第一幕",
    "act_2": "第二幕",
    "act_3": "第三幕",
    "chapter_outline": "章节细纲",
    "chapter": "章节",
    "act": "所属幕",
    "summary": "内容概要",
}


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/api/story-jobs", response_model=StoryJobCreateResponse, status_code=202)
def create_story_job(request: StoryGenerationRequest) -> StoryJobCreateResponse:
    payload = _model_to_payload(request)
    job_id = uuid4().hex
    now = _now_iso()

    with jobs_lock:
        jobs[job_id] = {
            "job_id": job_id,
            "status": "pending",
            "created_at": now,
            "updated_at": now,
            "request": payload,
            "project_id": payload.get("project_id"),
            "result_state": None,
            "error": None,
        }

    job_executor.submit(_run_story_job, job_id, payload)

    return StoryJobCreateResponse(
        job_id=job_id,
        status="pending",
        status_url=f"/api/story-jobs/{job_id}",
        result_url=f"/api/story-jobs/{job_id}/result",
    )


@app.get("/api/story-jobs/{job_id}", response_model=StoryJobStatusResponse)
def get_story_job(job_id: str) -> StoryJobStatusResponse:
    job = _get_job(job_id)
    return _job_status_response(job)


@app.get("/api/story-jobs/{job_id}/result", response_model=StoryJobResultResponse)
def get_story_job_result(job_id: str) -> StoryJobResultResponse:
    job = _get_job(job_id)
    result = None
    if job["status"] == "succeeded" and job.get("result_state"):
        result = response_from_state(job["result_state"])

    return StoryJobResultResponse(
        job_id=job["job_id"],
        status=job["status"],
        result=result,
        error=job.get("error"),
    )


@app.get("/api/story-jobs/{job_id}/characters", response_model=CharactersResponse)
def get_story_job_characters(job_id: str) -> CharactersResponse:
    job = _get_job(job_id)
    if job["status"] != "succeeded":
        raise HTTPException(status_code=409, detail=f"Job is not completed yet: {job['status']}")

    state = job.get("result_state") or {}
    return _characters_response_from_state(
        state,
        job_id=job["job_id"],
        project_id=state.get("project_id") or job.get("project_id"),
        status=job["status"],
        file_url=f"/api/story-jobs/{job['job_id']}/files/characters",
    )


@app.get("/api/projects/{project_id}/characters", response_model=CharactersResponse)
def get_project_characters(project_id: str) -> CharactersResponse:
    path = _resolve_project_file(project_id, "characters")
    if path is None:
        raise HTTPException(status_code=404, detail=f"Characters file not found for project: {project_id}")

    characters = path.read_text(encoding="utf-8")
    parsed = parse_json_object(characters)
    return CharactersResponse(
        project_id=project_id,
        characters=characters,
        display_text=_display_text_from_content(characters, "characters", parsed),
        parsed=parsed,
        file_url=f"/api/projects/{project_id}/characters",
    )


@app.get("/api/story-jobs/{job_id}/relationships", response_model=StageSectionResponse)
def get_story_job_relationships(job_id: str) -> StageSectionResponse:
    return _story_job_stage_section_response(job_id, "relationships")


@app.get("/api/projects/{project_id}/relationships", response_model=StageSectionResponse)
def get_project_relationships(project_id: str) -> StageSectionResponse:
    return _project_stage_section_response(project_id, "relationships")


@app.get("/api/story-jobs/{job_id}/biography", response_model=StageSectionResponse)
def get_story_job_biography(job_id: str) -> StageSectionResponse:
    return _story_job_stage_section_response(job_id, "biography")


@app.get("/api/projects/{project_id}/biography", response_model=StageSectionResponse)
def get_project_biography(project_id: str) -> StageSectionResponse:
    return _project_stage_section_response(project_id, "biography")


@app.get("/api/story-jobs/{job_id}/outline", response_model=StageSectionResponse)
def get_story_job_outline(job_id: str) -> StageSectionResponse:
    return _story_job_stage_section_response(job_id, "outline")


@app.get("/api/projects/{project_id}/outline", response_model=StageSectionResponse)
def get_project_outline(project_id: str) -> StageSectionResponse:
    return _project_stage_section_response(project_id, "outline")


@app.get("/api/projects/{project_id}/storyboard", response_model=ProjectStoryboardResponse)
def get_project_storyboard(project_id: str) -> ProjectStoryboardResponse:
    path = _resolve_project_file(project_id, "storyboard")
    if path is None:
        raise HTTPException(status_code=404, detail=f"Storyboard file not found for project: {project_id}")

    rows = _read_storyboard_xlsx(path)
    return ProjectStoryboardResponse(
        project_id=project_id,
        rows=rows,
        display_text=_storyboard_rows_to_markdown(rows),
        file_url=f"/api/projects/{project_id}/files/storyboard",
    )


@app.get("/api/projects/{project_id}/files/{file_type}")
def download_project_file(project_id: str, file_type: str) -> FileResponse:
    file_key = FILE_TYPE_ALIASES.get(file_type)
    if file_key is None:
        supported = ", ".join(sorted(FILE_TYPE_ALIASES))
        raise HTTPException(status_code=400, detail=f"Unsupported file_type. Supported values: {supported}")

    file_path = _resolve_project_file(project_id, file_key)
    if file_path is None:
        raise HTTPException(status_code=404, detail=f"File is not available for type: {file_type}")

    return FileResponse(path=file_path, filename=file_path.name, media_type=_media_type_for(file_path))


@app.get("/api/story-jobs/{job_id}/files/{file_type}")
def download_story_job_file(job_id: str, file_type: str) -> FileResponse:
    job = _get_job(job_id)
    if job["status"] != "succeeded":
        raise HTTPException(status_code=409, detail=f"Job is not completed yet: {job['status']}")

    state = job.get("result_state") or {}
    file_key = FILE_TYPE_ALIASES.get(file_type)
    if file_key is None:
        supported = ", ".join(sorted(FILE_TYPE_ALIASES))
        raise HTTPException(status_code=400, detail=f"Unsupported file_type. Supported values: {supported}")

    file_path = _resolve_result_file(state, file_key)
    if file_path is None:
        raise HTTPException(status_code=404, detail=f"File is not available for type: {file_type}")

    media_type = _media_type_for(file_path)
    return FileResponse(path=file_path, filename=file_path.name, media_type=media_type)


@app.post("/generate", response_model=StoryGenerationResponse)
def generate_story(request: StoryGenerationRequest) -> StoryGenerationResponse:
    payload = _model_to_payload(request)

    try:
        state = orchestrator.generate(payload, include_storyboard=payload.get("include_storyboard", True))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return response_from_state(state)


@app.post("/generate/script-only", response_model=StoryGenerationResponse)
def generate_script_only(request: StoryGenerationRequest) -> StoryGenerationResponse:
    payload = _model_to_payload(request)

    try:
        state = orchestrator.generate(payload, include_storyboard=False)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return response_from_state(state)


@app.post("/storyboard", response_model=StoryGenerationResponse)
def generate_storyboard(request: StoryboardGenerationRequest) -> StoryGenerationResponse:
    payload = _model_to_payload(request)

    final_script = (payload.get("final_script") or "").strip()
    script_path = payload.get("script_path")
    project_id = payload.get("project_id")
    output_dir = payload.get("output_dir")

    try:
        if not final_script:
            if script_path:
                path = Path(script_path)
            elif project_id:
                path = orchestrator.settings.output_root / project_id / "final_script.md"
                script_path = str(path)
            else:
                raise ValueError("Provide final_script, script_path, or project_id.")

            if not path.exists():
                raise FileNotFoundError(f"final_script not found: {path}")
            final_script = path.read_text(encoding="utf-8")
            output_dir = output_dir or str(path.resolve().parent)

        state = orchestrator.generate_storyboard_from_script(
            final_script=final_script,
            project_id=project_id,
            output_dir=output_dir,
            final_script_path=script_path,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return response_from_state(state)


@app.on_event("shutdown")
def shutdown_job_executor() -> None:
    job_executor.shutdown(wait=False, cancel_futures=True)


def _run_story_job(job_id: str, payload: dict[str, Any]) -> None:
    _update_job(job_id, status="running", error=None)
    try:
        state = orchestrator.generate(payload, include_storyboard=payload.get("include_storyboard", True))
    except Exception as exc:
        _update_job(job_id, status="failed", error=str(exc))
        return

    _update_job(
        job_id,
        status="succeeded",
        project_id=state.get("project_id"),
        result_state=dict(state),
        error=None,
    )


def _get_job(job_id: str) -> dict[str, Any]:
    with jobs_lock:
        job = jobs.get(job_id)
        if job is None:
            raise HTTPException(status_code=404, detail=f"Job not found: {job_id}")
        return dict(job)


def _update_job(job_id: str, **updates: Any) -> None:
    with jobs_lock:
        job = jobs.get(job_id)
        if job is None:
            return
        job.update(updates)
        job["updated_at"] = _now_iso()


def _job_status_response(job: dict[str, Any]) -> StoryJobStatusResponse:
    return StoryJobStatusResponse(
        job_id=job["job_id"],
        status=job["status"],
        created_at=job["created_at"],
        updated_at=job["updated_at"],
        project_id=job.get("project_id"),
        error=job.get("error"),
        result_url=f"/api/story-jobs/{job['job_id']}/result",
        files=_available_file_urls(job),
    )


def _available_file_urls(job: dict[str, Any]) -> dict[str, str]:
    if job["status"] != "succeeded":
        return {}

    state = job.get("result_state") or {}
    stage_files = state.get("stage_files", {})
    urls = {}
    for file_key in stage_files:
        if file_key in FALLBACK_FILENAMES:
            urls[file_key] = f"/api/story-jobs/{job['job_id']}/files/{file_key}"
    return urls


def _resolve_result_file(state: dict[str, Any], file_key: str) -> Path | None:
    stage_files = state.get("stage_files", {})
    stage_path = stage_files.get(file_key)
    if stage_path:
        path = Path(stage_path)
        if path.exists() and path.is_file():
            return path

    output_dir = state.get("output_dir")
    fallback_filename = FALLBACK_FILENAMES.get(file_key)
    if output_dir and fallback_filename:
        path = Path(output_dir) / fallback_filename
        if path.exists() and path.is_file():
            return path
    return None


def _resolve_project_file(project_id: str, file_key: str) -> Path | None:
    filename = FALLBACK_FILENAMES.get(file_key)
    if filename is None:
        return None

    output_root = orchestrator.settings.output_root.resolve()
    path = (output_root / project_id / filename).resolve()
    try:
        path.relative_to(output_root)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid project_id")

    if path.exists() and path.is_file():
        return path
    return None


def _characters_response_from_state(
    state: dict[str, Any],
    *,
    job_id: str | None = None,
    project_id: str | None = None,
    status: JobStatus | None = None,
    file_url: str | None = None,
) -> CharactersResponse:
    characters = state.get("characters") or ""
    file_path = _resolve_result_file(state, "characters")

    if file_path is not None:
        if not characters:
            characters = file_path.read_text(encoding="utf-8")

    if not characters:
        raise HTTPException(status_code=404, detail="Characters content is not available.")

    parsed = parse_json_object(characters)
    return CharactersResponse(
        job_id=job_id,
        project_id=project_id,
        status=status,
        characters=characters,
        display_text=_display_text_from_content(characters, "characters", parsed),
        parsed=parsed,
        file_url=file_url,
    )


def _story_job_stage_section_response(job_id: str, section: str) -> StageSectionResponse:
    job = _get_job(job_id)
    if job["status"] != "succeeded":
        raise HTTPException(status_code=409, detail=f"Job is not completed yet: {job['status']}")

    state = job.get("result_state") or {}
    return _stage_section_response_from_state(
        state,
        section,
        job_id=job["job_id"],
        project_id=state.get("project_id") or job.get("project_id"),
        status=job["status"],
        file_url=f"/api/story-jobs/{job['job_id']}/files/{STAGE_FILE_KEYS[section]}",
    )


def _project_stage_section_response(project_id: str, section: str) -> StageSectionResponse:
    file_key = STAGE_FILE_KEYS[section]
    path = _resolve_project_file(project_id, file_key)
    if path is None:
        raise HTTPException(status_code=404, detail=f"{section} file not found for project: {project_id}")

    content = path.read_text(encoding="utf-8")
    return _build_stage_section_response(
        section=section,
        content=content,
        project_id=project_id,
        file_url=f"/api/projects/{project_id}/{section}",
    )


def _stage_section_response_from_state(
    state: dict[str, Any],
    section: str,
    *,
    job_id: str | None = None,
    project_id: str | None = None,
    status: JobStatus | None = None,
    file_url: str | None = None,
) -> StageSectionResponse:
    state_key = STAGE_SECTION_KEYS[section]
    file_key = STAGE_FILE_KEYS[section]
    content = state.get(state_key) or ""
    file_path = _resolve_result_file(state, file_key)

    if file_path is not None and not content:
        content = file_path.read_text(encoding="utf-8")

    if not content:
        raise HTTPException(status_code=404, detail=f"{section} content is not available.")

    return _build_stage_section_response(
        section=section,
        content=content,
        job_id=job_id,
        project_id=project_id,
        status=status,
        file_url=file_url,
    )


def _build_stage_section_response(
    *,
    section: str,
    content: str,
    job_id: str | None = None,
    project_id: str | None = None,
    status: JobStatus | None = None,
    file_url: str | None = None,
) -> StageSectionResponse:
    parsed = parse_json_object(content)
    values: dict[str, Any] = {
        "job_id": job_id,
        "project_id": project_id,
        "status": status,
        "section": section,
        "content": content,
        "display_text": _display_text_from_content(content, section, parsed),
        "parsed": parsed,
        "file_url": file_url,
    }
    values[section] = content
    return StageSectionResponse(**values)


def _display_text_from_content(content: str, section: str, parsed: Any = None) -> str:
    data = parsed if parsed is not None else parse_json_object(content)
    if data is None:
        return content.strip()

    if section == "characters":
        return _format_characters_display(data)
    if section == "relationships":
        return _format_relationships_display(data)
    if section == "biography":
        return _format_biography_display(data)
    if section == "outline":
        return _format_outline_display(data)
    return _format_display_value(data)


def _format_characters_display(data: Any) -> str:
    characters = _unwrap_section(data, "characters")
    if not isinstance(characters, dict):
        return _format_display_value(data)

    lines = ["# 角色设置"]
    for role_key, value in characters.items():
        lines.extend(["", f"## {_display_label(role_key)}"])
        lines.extend(_format_value_lines(value))
    return "\n".join(lines).strip()


def _format_relationships_display(data: Any) -> str:
    relationships = _unwrap_section(data, "relationships")
    if not isinstance(relationships, list):
        return _format_display_value(data)

    lines = ["# 人物关系"]
    for index, item in enumerate(relationships, start=1):
        if isinstance(item, dict):
            title = item.get("name") or f"关系 {index}"
            lines.extend(["", f"## {index}. {title}"])
            lines.extend(_format_mapping_lines(item, skip_keys={"name"}))
        else:
            lines.extend(["", f"## {index}. 关系", str(item)])
    return "\n".join(lines).strip()


def _format_biography_display(data: Any) -> str:
    if not isinstance(data, dict):
        return _format_display_value(data)

    lines = ["# 人物小传"]
    if data.get("protagonist_biography"):
        lines.extend(["", "## 主角小传", str(data["protagonist_biography"])])

    for key in ("core_supporting_biographies", "functional_supporting_biographies"):
        values = data.get(key)
        if values is None:
            continue
        lines.extend(["", f"## {_display_label(key)}"])
        lines.extend(_format_value_lines(values))

    handled = {"protagonist_biography", "core_supporting_biographies", "functional_supporting_biographies"}
    for key, value in data.items():
        if key not in handled:
            lines.extend(["", f"## {_display_label(key)}"])
            lines.extend(_format_value_lines(value))

    return "\n".join(lines).strip()


def _format_outline_display(data: Any) -> str:
    if not isinstance(data, dict):
        return _format_display_value(data)

    lines = ["# 故事大纲"]
    if data.get("duration_minutes") is not None:
        lines.extend(["", f"影片时长：{data['duration_minutes']} 分钟"])

    three_act = data.get("three_act_outline")
    if isinstance(three_act, dict):
        lines.extend(["", "## 三幕式大纲"])
        for key, value in three_act.items():
            lines.extend(["", f"### {_display_label(key)}"])
            lines.extend(_format_value_lines(value))

    chapters = data.get("chapter_outline")
    if isinstance(chapters, list):
        lines.extend(["", "## 章节细纲"])
        for index, chapter in enumerate(chapters, start=1):
            if isinstance(chapter, dict):
                title = chapter.get("chapter") or f"章节 {index}"
                lines.extend(["", f"### {index}. {title}"])
                lines.extend(_format_mapping_lines(chapter, skip_keys={"chapter"}))
            else:
                lines.extend(["", f"### {index}. 章节", str(chapter)])

    handled = {"duration_minutes", "three_act_outline", "chapter_outline"}
    for key, value in data.items():
        if key not in handled:
            lines.extend(["", f"## {_display_label(key)}"])
            lines.extend(_format_value_lines(value))

    return "\n".join(lines).strip()


def _format_display_value(value: Any) -> str:
    return "\n".join(_format_value_lines(value)).strip()


def _read_storyboard_xlsx(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on local env
        raise RuntimeError("Package 'openpyxl' is missing. Run: pip install -r requirements.txt") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[list[str]] = []

    for row in worksheet.iter_rows(values_only=True):
        values = ["" if cell is None else str(cell).strip() for cell in row]
        while values and values[-1] == "":
            values.pop()
        if values:
            rows.append(values)

    workbook.close()
    return rows


def _storyboard_rows_to_markdown(rows: list[list[str]]) -> str:
    if not rows:
        return ""

    width = max(len(row) for row in rows)

    def normalize(row: list[str]) -> list[str]:
        values = row[:]
        if len(values) < width:
            values.extend([""] * (width - len(values)))
        return [value.replace("|", "\\|").replace("\n", " ").strip() for value in values]

    header = normalize(rows[0])
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in rows[1:]:
        lines.append("| " + " | ".join(normalize(row)) + " |")
    return "\n".join(lines)


def _format_value_lines(value: Any) -> list[str]:
    if isinstance(value, dict):
        return _format_mapping_lines(value)
    if isinstance(value, list):
        lines: list[str] = []
        for index, item in enumerate(value, start=1):
            if isinstance(item, dict):
                title = item.get("name") or item.get("chapter") or f"第 {index} 项"
                lines.append(f"{index}. {title}")
                lines.extend(f"   {line}" for line in _format_mapping_lines(item, skip_keys={"name", "chapter"}))
            else:
                lines.append(f"{index}. {item}")
        return lines
    return [str(value)]


def _format_mapping_lines(mapping: dict[str, Any], skip_keys: set[str] | None = None) -> list[str]:
    skip_keys = skip_keys or set()
    lines: list[str] = []
    for key, value in mapping.items():
        if key in skip_keys:
            continue
        label = _display_label(key)
        if isinstance(value, (dict, list)):
            lines.append(f"{label}：")
            lines.extend(f"  {line}" for line in _format_value_lines(value))
        else:
            lines.append(f"{label}：{value}")
    return lines


def _unwrap_section(data: Any, key: str) -> Any:
    if isinstance(data, dict) and key in data:
        return data[key]
    return data


def _display_label(key: Any) -> str:
    return DISPLAY_LABELS.get(str(key), str(key))


def _media_type_for(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if suffix == ".json":
        return "application/json"
    if suffix in {".md", ".txt"}:
        return "text/plain; charset=utf-8"
    return "application/octet-stream"


def _model_to_payload(model: BaseModel) -> dict[str, Any]:
    try:
        return model.model_dump()
    except AttributeError:  # pydantic v1 fallback
        return model.dict()


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()
